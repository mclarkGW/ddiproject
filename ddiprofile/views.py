from django.db.models import Sum, F, FloatField, Count, Value
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models.functions import Lower
from .models import Initiative,ChangeRequest,LastRefreshed, Iteration, crChangeRequest,  crLastRefreshed, WBSInformation, gapLastRefreshed, gapChangeRequest  
from datetime import datetime, date
from django.core.management import call_command
from django.core.management.base import CommandError
from django.http import JsonResponse, HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill  
import requests
import base64
import json
from itertools import groupby
from collections import Counter, defaultdict


def index(request):
    return render(request, 'index.html')

# Home View
def home(request):
    return render(request, 'home.html')

# DDI Dashboard View
def dashboard_view(request):
    last_refreshed = LastRefreshed.objects.first()
    initiatives = Initiative.objects.all() \
        .prefetch_related(
            'change_requests'
            ) \
            .order_by('clientaccounts')
    
    # Annotate "Total_Planned" for each ChangeRequest
    cr = ChangeRequest.objects.all() \
        .prefetch_related('features', 'features__user_stories') \
        .annotate(
            total_planned=Sum(
                F('features__user_stories__storypoints') * 8,
                output_field=FloatField()
            )
        ) \
        .order_by('tt_workdescription')

    context = {
        'initiatives': initiatives,
        'cr': cr,
        'last_refreshed': last_refreshed,
        'active_tab': 'dashboard',
    }

    return render(request, 'dashboard.html', context)

# DDI Export to Excel
def export_ddi_dashboard(request):
    initiatives = Initiative.objects.all().order_by('clientaccounts').prefetch_related(
        'change_requests',
        'change_requests__features',
        'change_requests__features__user_stories'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Initiatives"

    # Header
    headers = [
        "Initiative ID", "Initiative Title", "Client Accounts", "Current Phase",
        "CR/GAP ID", "CR/GAP Title", "CR/GAP State",
        "Feature ID", "Feature Title", "Feature State",
        "User Story ID", "User Story Title", "User Story State"
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Set column widths
    column_widths = [18, 30, 20, 18, 18, 30, 14, 14, 30, 14, 14, 30, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    row_num = 2
    rows_to_hide = []
    for initiative in initiatives:
        ws.append([
            initiative.id, initiative.title, initiative.clientaccounts, initiative.currentphase,
            "", "", "",
            "", "", "",
            "", "", ""
        ])
        row_num += 1
        for cr in initiative.change_requests.all():
            ws.append([
                "", "", "", "",
                cr.id, cr.title, cr.state,
                "", "", "",
                "", "", ""
            ])
            ws.row_dimensions[row_num].outlineLevel = 1
            rows_to_hide.append(row_num)  # Hide by default
            row_num += 1
            for feature in cr.features.all():
                ws.append([
                    "", "", "", "",
                    "", "", "",
                    feature.id, feature.title, feature.state,
                    "", "", ""
                ])
                ws.row_dimensions[row_num].outlineLevel = 2
                rows_to_hide.append(row_num)
                row_num += 1
                for us in feature.user_stories.all():
                    ws.append([
                        "", "", "", "",
                        "", "", "",
                        "", "", "",
                        us.id, us.title, us.state
                    ])
                    ws.row_dimensions[row_num].outlineLevel = 3
                    rows_to_hide.append(row_num)
                    row_num += 1

    # Hide all grouped rows, so only top-level (Initiatives) show
    for rn in rows_to_hide:
        ws.row_dimensions[rn].hidden = True

    # Outline/summary settings
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_view.showOutlineSymbols = True

    # Add filter to header row
    ws.auto_filter.ref = "A1:M1"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=DDI_Dashboard_ALL_Items.xlsx'
    wb.save(response)
    return response

# Iteration View
def iteration_view(request):
    # Last Refresh Date
    iteration_last_refreshed = Iteration.objects.first()
    # Get all iterations ordered by path and name
    iterations = Iteration.objects.all().order_by('id')

    # Group iterations by release (Level 2)
    grouped_iterations = {}
    for iteration in iterations:
        # Split the path to extract the release level
        path_parts = iteration.path.split('\\')
        if len(path_parts) > 1:  # Ensure there is a second level (Release)
            release = path_parts[1]
        else:
            release = "Project Level"  # Default group for invalid paths

        if release not in grouped_iterations:
            grouped_iterations[release] = []
        grouped_iterations[release].append(iteration)

    context = {
        'iterations': grouped_iterations,
        'iteration_last_refreshed': iteration_last_refreshed,
    }
    return render(request, 'iteration.html', context)

# Fetch Iterations View
def run_fetch_iterations(request):
    last_refreshed = Iteration.objects.order_by('-scriptupdated').first()
    if request.method == "POST":
        try:
            # Call the custom management command
            call_command('fetch_iterations')
            return JsonResponse({"status": "success", "message": "Iterations fetched successfully!"})
        except CommandError as e:
            return JsonResponse({"status": "error", "message": str(e)})
    else:
        # Render the HTML template for GET requests
        return render(request, "run_fetch_iterations.html", {"last_refreshed": last_refreshed})

# Fetch DDI Data View
def run_fetch_data(request):
    last_refreshed = LastRefreshed.objects.first()
    if request.method == "POST":
        try:
            # Call the custom management command
            call_command('fetch_data')
            return JsonResponse({"status": "success", "message": "Updated Data fetched successfully!"})
        except CommandError as e:
            return JsonResponse({"status": "error", "message": str(e)})
    else:
        # Render the HTML template for GET requests
        return render(request, "run_fetch_data.html",{"last_refreshed": last_refreshed})

# Fetch CR Data View
def run_crfetch_data(request):
    last_refreshed = crLastRefreshed.objects.first()
    if request.method == "POST":
        try:
            # Call the custom management command
            call_command('crfetch_data')
            return JsonResponse({"status": "success", "message": "Updated Data fetched successfully!"})
        except CommandError as e:
            return JsonResponse({"status": "error", "message": str(e)})
    else:
        # Render the HTML template for GET requests
        return render(request, "run_crfetch_data.html",{"last_refreshed": last_refreshed})
    
# CR Dashboard View
def cr_dashboard_view(request):
    last_refreshed = crLastRefreshed.objects.first()

    # Build rollup: workdescription -> total effort
    wbs_effort_map = defaultdict(float)
    for wbs in WBSInformation.objects.values('workdescription', 'effort'):
        wd = wbs['workdescription']
        effort = wbs['effort'] or 0
        if wd:
            wbs_effort_map[wd] += float(effort)

    # Prefetch features and their user stories for optimized queries
    cr_qs = crChangeRequest.objects.all() \
        .prefetch_related(
            'features',
            'features__user_stories',
            'features__user_stories__tasks'
        ) \
        .order_by('clientaccounts')
    cr = list(cr_qs)  # We'll be working with an in-memory list for easier filtering

    # Build the set of all modules for the dropdown
    all_modules = set()
    for change_request in cr:
        for feature in change_request.features.all():
            if feature.tt_initiative:
                all_modules.add(feature.tt_initiative)
    all_modules = sorted(all_modules)

    # Build the set of all client accounts for the dropdown
    all_client_accounts = set()
    for change_request in cr:
        ca = getattr(change_request, "clientaccounts", None)
        if ca and ca.strip():
            # If clientaccounts is a semicolon-separated list, split and trim:
            for acc in str(ca).split(";"):
                acc = acc.strip()
                if acc:
                    all_client_accounts.add(acc)
    all_client_accounts = sorted(all_client_accounts)

    # Get selected filters from GET params
    selected_state = request.GET.get("cr_state", "")
    selected_modules = request.GET.getlist("module")
    if "__all__" in selected_modules or not selected_modules:
        selected_modules = []  # Treat as "no filter"
    else:
        cr = [
            c for c in cr
            if any(f.tt_initiative in selected_modules for f in c.features.all())
        ]
    selected_client_accounts = request.GET.getlist("client_account")
    if "__all__" in selected_client_accounts or not selected_client_accounts:
        selected_client_accounts = []  # Treat as "no filter"
    else:
        # Apply filter to CR list
        cr = [
            c for c in cr
            if any(acc in selected_client_accounts for acc in str(getattr(c, "clientaccounts", "")).split(";"))
        ]

    # Apply state filter
    if selected_state == "HideDone":
        cr = [c for c in cr if (c.state or "").strip().lower() != "done"]
    elif selected_state == "InProgress":
        progress_states = ["committed", "in progress", "blocked"]
        cr = [c for c in cr if (c.state or "").strip().lower() in progress_states]
    elif selected_state == "Proposed":
        proposed_states = ["funnel", "reviewing", "analyzing", "backlog"]
        cr = [c for c in cr if (c.state or "").strip().lower() in proposed_states]
    # Else: show all

    # Apply module filter
    if selected_modules:
        cr = [
            c for c in cr
            if any(f.tt_initiative in selected_modules for f in c.features.all())
        ]

    # Now rollups and calculations (only for filtered CRs)
    for change_request in cr:
        # 1. Story Point Rollups
        total_planned = 0
        total_done = 0
        total_active = 0
        total_new = 0

        # 2. PI Start/End across all features (for CR)
        cr_pi_start = None
        cr_pi_end = None

        for feature in change_request.features.all():
            # 3. PI Start/End for Feature
            pi_start = None
            pi_end = None

            for user_story in feature.user_stories.all():
                # Story Points Rollup
                try:
                    storypoints = float(user_story.storypoints)
                except (ValueError, TypeError):
                    storypoints = 0

                state = getattr(user_story, "state", "").strip().lower()
                # Planned (all user stories)
                total_planned += storypoints * 8
                # Done/Active/New
                if state == "done":
                    total_done += storypoints
                elif state in ("ready", "in progress", "test", "blocked"):
                    total_active += storypoints
                elif state in ("new", "backlog"):
                    total_new += storypoints

                # PI Start/End per Feature
                iteration_path = getattr(user_story, "iterationpath", None)
                if iteration_path:
                    try:
                        iteration = Iteration.objects.get(path=iteration_path)
                        if not pi_start or (iteration.start_date and iteration.start_date < pi_start):
                            pi_start = iteration.start_date
                        if not pi_end or (iteration.finish_date and iteration.finish_date > pi_end):
                            pi_end = iteration.finish_date
                    except Iteration.DoesNotExist:
                        pass

            # Attach feature-level PI dates if you want them on the feature:
            feature.pi_start = pi_start
            feature.pi_end = pi_end

            # CR-level PI dates: accumulate across all features
            if pi_start:
                if not cr_pi_start or pi_start < cr_pi_start:
                    cr_pi_start = pi_start
            if pi_end:
                if not cr_pi_end or pi_end > cr_pi_end:
                    cr_pi_end = pi_end

        # Attach rollups to the CR
        change_request.total_planned = total_planned
        change_request.total_done = total_done
        change_request.total_active = total_active
        change_request.total_new = total_new
        change_request.pi_start = cr_pi_start
        change_request.pi_end = cr_pi_end
        change_request.total_sp = total_done + total_active + total_new  # Total story points (in "hours") as sum of all states

        # Progress percentage for progress bar
        if change_request.total_sp > 0:
            change_request.percent_done = int(round((change_request.total_done / change_request.total_sp) * 100))
        else:
            change_request.percent_done = 0

        # +/- Target Date calculation (CR level)
        target = getattr(change_request, 'targetdate', None)
        pi_end = cr_pi_end
        if target and pi_end:
            days_diff = (target - pi_end).days
            if days_diff > 0:
                change_request.target_vs_pi = f'+{days_diff}'
            elif days_diff < 0:
                change_request.target_vs_pi = f'{days_diff}'
            else:
                change_request.target_vs_pi = '0'
        else:
            change_request.target_vs_pi = ''

        # Attach WBS effort rollup to each CR using tt_workdescription
        workdesc = getattr(change_request, 'tt_workdescription', None)
        change_request.wbs_total_effort = wbs_effort_map.get(workdesc, 0)

        # --- TASK ROLLUPS ---
        total_originalestimate = 0.0
        total_remainingwork = 0.0
        total_completedwork = 0.0
        all_tasks = []

        for feature in change_request.features.all():
            for user_story in feature.user_stories.all():
                for task in user_story.tasks.all():
                    all_tasks.append(task)
                    # originalestimate
                    try:
                        total_originalestimate += float(task.originalestimate or 0)
                    except (ValueError, TypeError):
                        pass
                    # remainingwork
                    try:
                        total_remainingwork += float(task.remainingwork or 0)
                    except (ValueError, TypeError):
                        pass
                    # completedwork
                    try:
                        total_completedwork += float(task.completedwork or 0)
                    except (ValueError, TypeError):
                        pass

        change_request.tasks = all_tasks  # List of all tasks under this CR (for display)
        change_request.total_task_originalestimate = total_originalestimate
        change_request.total_task_remainingwork = total_remainingwork
        change_request.total_task_completedwork = total_completedwork

        # *** Planned vs Estimate Calculation ***
        highlevelestimate_raw = getattr(change_request, 'highlevelestimate', 0) or 0
        try:
            highlevelestimate = float(highlevelestimate_raw)
        except (ValueError, TypeError):
            highlevelestimate = 0
        change_request.planned_vs_estimate = change_request.total_task_originalestimate - highlevelestimate

    context = {
        'last_refreshed': last_refreshed,
        'cr': cr,
        'all_modules': all_modules,
        'selected_modules': selected_modules,
        'selected_state': selected_state,
        'all_client_accounts': all_client_accounts,
        'selected_client_accounts': selected_client_accounts,
        'active_tab': 'dashboard',
    }
    return render(request, 'crdashboard.html', context)

# CR Export to Excel
def export_cr_dashboard(request):
    # Query all change requests with features, user stories, and tasks
    cr_qs = crChangeRequest.objects.all().order_by('clientaccounts').prefetch_related(
        'features',
        'features__user_stories',
        'features__user_stories__tasks'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Change Requests"

    # Header
    headers = [
        "CR ID", "CR Title", "CR State", "CR Workdesc", "CR High Level Estimate", "CR Client Accounts",
        "Feature ID", "Feature Title", "Feature State", "Feature Module",
        "User Story ID", "User Story Title", "User Story State", "User Story Story Points",
        "Task ID", "Task Title", "Task State", "Task Original Estimate", "Task Remaining Work", "Task Completed Work"
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Set column widths (adjust as needed)
    column_widths = [14, 30, 14, 18, 16, 20, 14, 30, 14, 16, 14, 30, 14, 10, 14, 30, 14, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    row_num = 2
    rows_to_hide = []
    for cr in cr_qs:
        ws.append([
            cr.id, cr.title, cr.state, cr.tt_workdescription, cr.highlevelestimate, cr.clientaccounts,
            "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])
        row_num += 1
        # Features
        for feature in cr.features.all():
            ws.append([
                "", "", "", "", "", "",
                feature.id, feature.title, feature.state, feature.tt_initiative,
                "", "", "", "", "", "", "", "", "", ""
            ])
            ws.row_dimensions[row_num].outlineLevel = 1
            rows_to_hide.append(row_num)
            row_num += 1
            # User Stories
            for us in feature.user_stories.all():
                ws.append([
                    "", "", "", "", "", "",
                    "", "", "", "",
                    us.id, us.title, us.state, us.storypoints,
                    "", "", "", "", "", ""
                ])
                ws.row_dimensions[row_num].outlineLevel = 2
                rows_to_hide.append(row_num)
                row_num += 1
                # Tasks (optional: comment out if not needed)
                for task in us.tasks.all():
                    ws.append([
                        "", "", "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        task.id, getattr(task, "title", ""), getattr(task, "state", ""),
                        getattr(task, "originalestimate", ""), getattr(task, "remainingwork", ""), getattr(task, "completedwork", "")
                    ])
                    ws.row_dimensions[row_num].outlineLevel = 3
                    rows_to_hide.append(row_num)
                    row_num += 1

    # Hide all grouped rows, so only top-level CRs show
    for rn in rows_to_hide:
        ws.row_dimensions[rn].hidden = True

    # Outline/summary settings
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_view.showOutlineSymbols = True

    # Add filter to header row
    ws.auto_filter.ref = "A1:T1"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=CR_Dashboard_ALL_Items.xlsx'
    wb.save(response)
    return response

# CR Charts View
def cr_charts_view(request):
    crs = crChangeRequest.objects.prefetch_related('features', 'features__user_stories')

    # RELEASE aggregation (already correct)
    last_release_list = []
    for cr in crs:
        all_userstory_releases = []
        for feature in cr.features.all():
            for user_story in feature.user_stories.all():
                userstory_release = getattr(user_story, "iterationpath", None)
                if userstory_release:
                    if "\\" in userstory_release:
                        userstory_release = userstory_release.split("\\", 1)[1]
                    all_userstory_releases.append(userstory_release)
        last_release = max(all_userstory_releases) if all_userstory_releases else None
        last_release_list.append(last_release if last_release else '(blank)')
    release_counter = Counter(last_release_list)
    sorted_release_items = sorted(release_counter.items(), key=lambda x: x[0])
    release_labels = [item[0] for item in sorted_release_items]
    release_data = [item[1] for item in sorted_release_items]

    # ACCOUNT aggregation (fix here)
    account_list = []
    for cr in crs:
        account = getattr(cr, 'clientaccounts', None)
        account_list.append(account if account else '(blank)')
    account_counter = Counter(account_list)
    sorted_account_items = sorted(account_counter.items(), key=lambda x: (x[0] == '(blank)', x[0]))
    account_labels = [item[0] for item in sorted_account_items]
    account_data = [item[1] for item in sorted_account_items]

    # STATUS aggregation (custom order)
    status_order = [
        'Funnel', 'Reviewing', 'Analyzing', 'Backlog',
        'Committed', 'In Progress', 'BLOCKED', 'Done'
    ]
    order_map = {name.lower(): i for i, name in enumerate(status_order)}
    status_list = []
    for cr in crs:
        status = getattr(cr, 'state', None)
        status_list.append(status if status else '(blank)')
    status_counter = Counter(status_list)
    sorted_status_items = sorted(
        status_counter.items(),
        key=lambda x: order_map.get(str(x[0]).strip().lower(), len(order_map))
    )
    status_labels = [item[0] for item in sorted_status_items]
    status_data = [item[1] for item in sorted_status_items]

    context = {
        'release_labels': json.dumps(release_labels),
        'release_data': json.dumps(release_data),
        'account_labels': json.dumps(account_labels),
        'account_data': json.dumps(account_data),
        'status_labels': json.dumps(status_labels),
        'status_data': json.dumps(status_data),
        'active_tab': 'charts',
    }
    return render(request, 'cr_charts.html', context)

# GAP Dashboard View
def gap_dashboard_view(request):
    last_refreshed = gapLastRefreshed.objects.first()

    # Build rollup: workdescription -> total effort
    wbs_effort_map = defaultdict(float)
    for wbs in WBSInformation.objects.values('workdescription', 'effort'):
        wd = wbs['workdescription']
        effort = wbs['effort'] or 0
        if wd:
            wbs_effort_map[wd] += float(effort)

    # Prefetch features and their user stories for optimized queries
    cr_qs = gapChangeRequest.objects.all() \
        .prefetch_related(
            'features',
            'features__user_stories',
            'features__user_stories__tasks'
        ) \
        .order_by('clientaccounts')
    cr = list(cr_qs)  # We'll be working with an in-memory list for easier filtering

    # Build the set of all modules for the dropdown
    all_modules = set()
    for change_request in cr:
        for feature in change_request.features.all():
            if feature.tt_initiative:
                all_modules.add(feature.tt_initiative)
    all_modules = sorted(all_modules)

    # Build the set of all client accounts for the dropdown
    all_client_accounts = set()
    for change_request in cr:
        ca = getattr(change_request, "clientaccounts", None)
        if ca and ca.strip():
            # If clientaccounts is a semicolon-separated list, split and trim:
            for acc in str(ca).split(";"):
                acc = acc.strip()
                if acc:
                    all_client_accounts.add(acc)
    all_client_accounts = sorted(all_client_accounts)

    # Get selected filters from GET params
    selected_state = request.GET.get("cr_state", "")
    selected_modules = request.GET.getlist("module")
    if "__all__" in selected_modules or not selected_modules:
        selected_modules = []  # Treat as "no filter"
    else:
        cr = [
            c for c in cr
            if any(f.tt_initiative in selected_modules for f in c.features.all())
        ]
    selected_client_accounts = request.GET.getlist("client_account")
    if "__all__" in selected_client_accounts or not selected_client_accounts:
        selected_client_accounts = []  # Treat as "no filter"
    else:
        # Apply filter to CR list
        cr = [
            c for c in cr
            if any(acc in selected_client_accounts for acc in str(getattr(c, "clientaccounts", "")).split(";"))
        ]

    # Apply state filter
    if selected_state == "HideDone":
        cr = [c for c in cr if (c.state or "").strip().lower() != "done"]
    elif selected_state == "InProgress":
        progress_states = ["committed", "in progress", "blocked"]
        cr = [c for c in cr if (c.state or "").strip().lower() in progress_states]
    elif selected_state == "Proposed":
        proposed_states = ["funnel", "reviewing", "analyzing", "backlog"]
        cr = [c for c in cr if (c.state or "").strip().lower() in proposed_states]
    # Else: show all

    # Apply module filter
    if selected_modules:
        cr = [
            c for c in cr
            if any(f.tt_initiative in selected_modules for f in c.features.all())
        ]

    # Now rollups and calculations (only for filtered CRs)
    for change_request in cr:
        # 1. Story Point Rollups
        total_planned = 0
        total_done = 0
        total_active = 0
        total_new = 0

        # 2. PI Start/End across all features (for CR)
        cr_pi_start = None
        cr_pi_end = None

        for feature in change_request.features.all():
            # 3. PI Start/End for Feature
            pi_start = None
            pi_end = None

            for user_story in feature.user_stories.all():
                # Story Points Rollup
                try:
                    storypoints = float(user_story.storypoints)
                except (ValueError, TypeError):
                    storypoints = 0

                state = getattr(user_story, "state", "").strip().lower()
                # Planned (all user stories)
                total_planned += storypoints * 8
                # Done/Active/New
                if state == "done":
                    total_done += storypoints
                elif state in ("ready", "in progress", "test", "blocked"):
                    total_active += storypoints
                elif state in ("new", "backlog"):
                    total_new += storypoints

                # PI Start/End per Feature
                iteration_path = getattr(user_story, "iterationpath", None)
                if iteration_path:
                    try:
                        iteration = Iteration.objects.get(path=iteration_path)
                        if not pi_start or (iteration.start_date and iteration.start_date < pi_start):
                            pi_start = iteration.start_date
                        if not pi_end or (iteration.finish_date and iteration.finish_date > pi_end):
                            pi_end = iteration.finish_date
                    except Iteration.DoesNotExist:
                        pass

            # Attach feature-level PI dates if you want them on the feature:
            feature.pi_start = pi_start
            feature.pi_end = pi_end

            # CR-level PI dates: accumulate across all features
            if pi_start:
                if not cr_pi_start or pi_start < cr_pi_start:
                    cr_pi_start = pi_start
            if pi_end:
                if not cr_pi_end or pi_end > cr_pi_end:
                    cr_pi_end = pi_end

        # Attach rollups to the CR
        change_request.total_planned = total_planned
        change_request.total_done = total_done
        change_request.total_active = total_active
        change_request.total_new = total_new
        change_request.pi_start = cr_pi_start
        change_request.pi_end = cr_pi_end
        change_request.total_sp = total_done + total_active + total_new  # Total story points (in "hours") as sum of all states

        # Progress percentage for progress bar
        if change_request.total_sp > 0:
            change_request.percent_done = int(round((change_request.total_done / change_request.total_sp) * 100))
        else:
            change_request.percent_done = 0

        # +/- Target Date calculation (CR level)
        target = getattr(change_request, 'targetdate', None)
        pi_end = cr_pi_end
        if target and pi_end:
            days_diff = (target - pi_end).days
            if days_diff > 0:
                change_request.target_vs_pi = f'+{days_diff}'
            elif days_diff < 0:
                change_request.target_vs_pi = f'{days_diff}'
            else:
                change_request.target_vs_pi = '0'
        else:
            change_request.target_vs_pi = ''

        # Attach WBS effort rollup to each CR using tt_workdescription
        workdesc = getattr(change_request, 'tt_workdescription', None)
        change_request.wbs_total_effort = wbs_effort_map.get(workdesc, 0)

        # --- TASK ROLLUPS ---
        total_originalestimate = 0.0
        total_remainingwork = 0.0
        total_completedwork = 0.0
        all_tasks = []

        for feature in change_request.features.all():
            for user_story in feature.user_stories.all():
                for task in user_story.tasks.all():
                    all_tasks.append(task)
                    # originalestimate
                    try:
                        total_originalestimate += float(task.originalestimate or 0)
                    except (ValueError, TypeError):
                        pass
                    # remainingwork
                    try:
                        total_remainingwork += float(task.remainingwork or 0)
                    except (ValueError, TypeError):
                        pass
                    # completedwork
                    try:
                        total_completedwork += float(task.completedwork or 0)
                    except (ValueError, TypeError):
                        pass

        change_request.tasks = all_tasks  # List of all tasks under this CR (for display)
        change_request.total_task_originalestimate = total_originalestimate
        change_request.total_task_remainingwork = total_remainingwork
        change_request.total_task_completedwork = total_completedwork

        # *** Planned vs Estimate Calculation ***
        highlevelestimate_raw = getattr(change_request, 'highlevelestimate', 0) or 0
        try:
            highlevelestimate = float(highlevelestimate_raw)
        except (ValueError, TypeError):
            highlevelestimate = 0
        change_request.planned_vs_estimate = change_request.total_task_originalestimate - highlevelestimate

    context = {
        'last_refreshed': last_refreshed,
        'cr': cr,
        'all_modules': all_modules,
        'selected_modules': selected_modules,
        'selected_state': selected_state,
        'all_client_accounts': all_client_accounts,
        'selected_client_accounts': selected_client_accounts,
        'active_tab': 'dashboard',
    }
    return render(request, 'gapdashboard.html', context)

# GAP Export to Exceldef export_gap_dashboard(request):
    gap_qs = gapChangeRequest.objects.all().prefetch_related(
        'features',
        'features__user_stories',
        'features__user_stories__tasks'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GAPs"

    # Header row
    headers = [
        "GAP ID", "GAP Title", "GAP State", "GAP Workdesc", "GAP High Level Estimate", "GAP Client Accounts",
        "Feature ID", "Feature Title", "Feature State", "Feature Module",
        "User Story ID", "User Story Title", "User Story State", "User Story Story Points",
        "Task ID", "Task Title", "Task State", "Task Original Estimate", "Task Remaining Work", "Task Completed Work"
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Set column widths (adjust as needed)
    column_widths = [14, 30, 14, 18, 16, 20, 14, 30, 14, 16, 14, 30, 14, 10, 14, 30, 14, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    row_num = 2
    rows_to_hide = []
    for gap in gap_qs:
        ws.append([
            gap.id, gap.title, gap.state, getattr(gap, 'tt_workdescription', ""), getattr(gap, 'highlevelestimate', ""), getattr(gap, 'clientaccounts', ""),
            "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])
        row_num += 1
        # Features
        for feature in gap.features.all():
            ws.append([
                "", "", "", "", "", "",
                feature.id, feature.title, feature.state, getattr(feature, 'tt_initiative', ""),
                "", "", "", "", "", "", "", "", "", ""
            ])
            ws.row_dimensions[row_num].outlineLevel = 1
            rows_to_hide.append(row_num)
            row_num += 1
            # User Stories
            for us in feature.user_stories.all():
                ws.append([
                    "", "", "", "", "", "",
                    "", "", "", "",
                    us.id, us.title, us.state, getattr(us, "storypoints", ""),
                    "", "", "", "", "", ""
                ])
                ws.row_dimensions[row_num].outlineLevel = 2
                rows_to_hide.append(row_num)
                row_num += 1
                # Tasks (optional)
                for task in us.tasks.all():
                    ws.append([
                        "", "", "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        task.id, getattr(task, "title", ""), getattr(task, "state", ""),
                        getattr(task, "originalestimate", ""), getattr(task, "remainingwork", ""), getattr(task, "completedwork", "")
                    ])
                    ws.row_dimensions[row_num].outlineLevel = 3
                    rows_to_hide.append(row_num)
                    row_num += 1

    # Hide all grouped rows, so only top-level GAPs show
    for rn in rows_to_hide:
        ws.row_dimensions[rn].hidden = True

    # Outline/summary settings
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_view.showOutlineSymbols = True

    # Add filter to header row
    ws.auto_filter.ref = "A1:T1"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=GAP_Dashboard_ALL_Items.xlsx'
    wb.save(response)
    return response

# GAP Export to Excel
def export_gap_dashboard(request):
    gap_qs = gapChangeRequest.objects.all().order_by('clientaccounts').prefetch_related(
        'features',
        'features__user_stories',
        'features__user_stories__tasks'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GAPs"

    # Header row
    headers = [
        "GAP ID", "GAP Title", "GAP State", "GAP Workdesc", "GAP High Level Estimate", "GAP Client Accounts",
        "Feature ID", "Feature Title", "Feature State", "Feature Module",
        "User Story ID", "User Story Title", "User Story State", "User Story Story Points",
        "Task ID", "Task Title", "Task State", "Task Original Estimate", "Task Remaining Work", "Task Completed Work"
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Set column widths (adjust as needed)
    column_widths = [14, 30, 14, 18, 16, 20, 14, 30, 14, 16, 14, 30, 14, 10, 14, 30, 14, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    row_num = 2
    rows_to_hide = []
    for gap in gap_qs:
        ws.append([
            gap.id, gap.title, gap.state, getattr(gap, 'tt_workdescription', ""), getattr(gap, 'highlevelestimate', ""), getattr(gap, 'clientaccounts', ""),
            "", "", "", "", "", "", "", "", "", "", "", "", "", ""
        ])
        row_num += 1
        # Features
        for feature in gap.features.all():
            ws.append([
                "", "", "", "", "", "",
                feature.id, feature.title, feature.state, getattr(feature, 'tt_initiative', ""),
                "", "", "", "", "", "", "", "", "", ""
            ])
            ws.row_dimensions[row_num].outlineLevel = 1
            rows_to_hide.append(row_num)
            row_num += 1
            # User Stories
            for us in feature.user_stories.all():
                ws.append([
                    "", "", "", "", "", "",
                    "", "", "", "",
                    us.id, us.title, us.state, getattr(us, "storypoints", ""),
                    "", "", "", "", "", ""
                ])
                ws.row_dimensions[row_num].outlineLevel = 2
                rows_to_hide.append(row_num)
                row_num += 1
                # Tasks (optional)
                for task in us.tasks.all():
                    ws.append([
                        "", "", "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        task.id, getattr(task, "title", ""), getattr(task, "state", ""),
                        getattr(task, "originalestimate", ""), getattr(task, "remainingwork", ""), getattr(task, "completedwork", "")
                    ])
                    ws.row_dimensions[row_num].outlineLevel = 3
                    rows_to_hide.append(row_num)
                    row_num += 1

    # Hide all grouped rows, so only top-level GAPs show
    for rn in rows_to_hide:
        ws.row_dimensions[rn].hidden = True

    # Outline/summary settings
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_view.showOutlineSymbols = True

    # Add filter to header row
    ws.auto_filter.ref = "A1:T1"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=GAP_Dashboard_ALL_Items.xlsx'
    wb.save(response)
    return response

# GAP Charts View
def gap_charts_view(request):
    crs = gapChangeRequest.objects.prefetch_related('features', 'features__user_stories')

    # RELEASE aggregation (already correct)
    last_release_list = []
    for cr in crs:
        all_userstory_releases = []
        for feature in cr.features.all():
            for user_story in feature.user_stories.all():
                userstory_release = getattr(user_story, "iterationpath", None)
                if userstory_release:
                    if "\\" in userstory_release:
                        userstory_release = userstory_release.split("\\", 1)[1]
                    all_userstory_releases.append(userstory_release)
        last_release = max(all_userstory_releases) if all_userstory_releases else None
        last_release_list.append(last_release if last_release else '(blank)')
    release_counter = Counter(last_release_list)
    sorted_release_items = sorted(release_counter.items(), key=lambda x: x[0])
    release_labels = [item[0] for item in sorted_release_items]
    release_data = [item[1] for item in sorted_release_items]

    # ACCOUNT aggregation (fix here)
    account_list = []
    for cr in crs:
        account = getattr(cr, 'clientaccounts', None)
        account_list.append(account if account else '(blank)')
    account_counter = Counter(account_list)
    sorted_account_items = sorted(account_counter.items(), key=lambda x: (x[0] == '(blank)', x[0]))
    account_labels = [item[0] for item in sorted_account_items]
    account_data = [item[1] for item in sorted_account_items]

    # STATUS aggregation (custom order)
    status_order = [
        'Funnel', 'Reviewing', 'Analyzing', 'Backlog',
        'Committed', 'In Progress', 'BLOCKED', 'Done'
    ]
    order_map = {name.lower(): i for i, name in enumerate(status_order)}
    status_list = []
    for cr in crs:
        status = getattr(cr, 'state', None)
        status_list.append(status if status else '(blank)')
    status_counter = Counter(status_list)
    sorted_status_items = sorted(
        status_counter.items(),
        key=lambda x: order_map.get(str(x[0]).strip().lower(), len(order_map))
    )
    status_labels = [item[0] for item in sorted_status_items]
    status_data = [item[1] for item in sorted_status_items]

    context = {
        'release_labels': json.dumps(release_labels),
        'release_data': json.dumps(release_data),
        'account_labels': json.dumps(account_labels),
        'account_data': json.dumps(account_data),
        'status_labels': json.dumps(status_labels),
        'status_data': json.dumps(status_data),
        'active_tab': 'charts',
    }
    return render(request, 'gap_charts.html', context)